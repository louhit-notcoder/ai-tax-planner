"""Excel (.xlsx, .xls) adapters for Indian tax documents.

Handles capital gains statements from brokers like Zerodha, Groww, etc.
Also supports bank statements in Excel format.
"""

from __future__ import annotations

import io
import re
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

from .base import AdapterResult, DocumentAdapter, ExtractedClaim, SourceLocation, money_value, parse_decimal

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _parse_decimal_excel(value: Any) -> Decimal | None:
    """Parse decimal from Excel cell value."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    text = str(value).strip().replace("₹", "").replace(",", "").replace(" ", "")
    text = re.sub(r"[A-Za-z$€£]", "", text)
    text = re.sub(r"[()]", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    if not text or text == "-":
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _parse_date_excel(value: Any) -> str | None:
    """Parse date from Excel cell value."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        # Excel date serial number (days since 1900-01-01)
        try:
            from datetime import date as date_class
            # Excel epoch is 1900-01-01 (with a bug for the 60th day)
            excel_epoch = date_class(1899, 12, 30)
            result = excel_epoch.replace(year=excel_epoch.year) + datetime.min.replace(hour=0, minute=0, second=0, microsecond=0).replace(hour=0) - datetime.min
            # Simpler approach
            d = datetime(1899, 12, 30) + timedelta(days=int(value))
            return d.date().isoformat()
        except Exception:
            return None
    text = str(value).strip()
    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d %b %Y", "%d-%b-%Y", "%Y/%m/%d"]:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _read_excel_sheet(sheet) -> list[dict[str, Any]]:
    """Read an Excel sheet into a list of dictionaries."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row (first row with text values)
    headers = []
    header_idx = 0
    for idx, row in enumerate(rows):
        # Check if row has any string values (likely headers)
        string_values = [str(v).strip() if v is not None else "" for v in row]
        if any(v for v in string_values if v):
            headers = string_values
            header_idx = idx
            break

    if not headers:
        return []

    # Normalize headers
    normalized_headers = []
    for h in headers:
        h_lower = str(h).lower().strip()
        # Common header mappings
        if any(k in h_lower for k in ["symbol", "scrip", "security", "scheme", "stock", "share"]):
            normalized_headers.append("Symbol")
        elif any(k in h_lower for k in ["buy", "acquisition", "purchase", "date"]):
            normalized_headers.append("Buy Date")
        elif any(k in h_lower for k in ["sell", "transfer", "sale", "disposal"]):
            normalized_headers.append("Sell Date")
        elif any(k in h_lower for k in ["sale", "proceeds", "sell value"]):
            normalized_headers.append("Sale Value")
        elif any(k in h_lower for k in ["buy", "cost", "purchase"]):
            normalized_headers.append("Buy Value")
        elif any(k in h_lower for k in ["expense", "brokerage", "stt", "charges"]):
            normalized_headers.append("Expenses")
        elif any(k in h_lower for k in ["description", "narration", "particular"]):
            normalized_headers.append("Description")
        else:
            normalized_headers.append(str(h).strip())

    # Read data rows
    results = []
    for row in rows[header_idx + 1:]:
        if not any(v for v in row if v is not None):
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(normalized_headers):
                row_dict[normalized_headers[i]] = val
        if row_dict:
            results.append(row_dict)

    return results


class ExcelCapitalGainsAdapter(DocumentAdapter):
    """Adapter for capital gains statements in Excel format.

    Supports exports from:
    - Zerodha
    - Groww
    - ICICI Direct
    - HDFC Securities
    - Angel Broking
    - And other Indian brokers
    """
    code = "EXCEL_CAPITAL_GAINS"
    version = "1.0.0"
    document_type = "BROKER_CAPITAL_GAINS"

    def supports(self, filename: str, mime_type: str, content: bytes) -> Decimal:
        if not HAS_OPENPYXL:
            return Decimal("0")
        name = filename.lower()
        if not (name.endswith(".xlsx") or name.endswith(".xls") or "spreadsheet" in (mime_type or "").lower()):
            return Decimal("0")
        # Check content for capital gains keywords
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text = " ".join(str(cell.value or "") for row in sheet.iter_rows(max_row=10, max_col=10) for cell in row)
                if any(k in text.lower() for k in ["capital gain", "capital loss", "sale proceeds", "buy", "sell", "equity", "mutual fund", "stocks", "shares", "trade"]):
                    return Decimal("0.85")
            wb.close()
        except Exception:
            pass
        return Decimal("0.25")

    def extract(self, filename: str, mime_type: str, content: bytes) -> AdapterResult:
        if not HAS_OPENPYXL:
            return AdapterResult(
                self.code, self.version, self.document_type,
                warnings=["openpyxl not installed. Please install: pip install openpyxl"]
            )

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            return AdapterResult(
                self.code, self.version, self.document_type,
                warnings=[f"Could not read Excel file: {str(e)}"]
            )

        claims: list[ExtractedClaim] = []
        warnings: list[str] = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = _read_excel_sheet(sheet)

            for idx, row in enumerate(rows):
                symbol = str(row.get("Symbol") or row.get("Description") or f"ASSET_{idx}").strip()

                buy_date = _parse_date_excel(row.get("Buy Date"))
                sell_date = _parse_date_excel(row.get("Sell Date"))

                sale = _parse_decimal_excel(row.get("Sale Value"))
                cost = _parse_decimal_excel(row.get("Buy Value"))
                expenses = _parse_decimal_excel(row.get("Expenses")) or Decimal("0")

                if sale is None and cost is None:
                    continue

                entity = f"TRADE_{idx}_{re.sub(r'[^a-zA-Z0-9]', '_', symbol[:30])}"[:160]

                # Determine asset type
                row_text = " ".join(str(v) for v in row.values() if v).lower()
                if any(k in row_text for k in ["mutual fund", "mf", "scheme"]):
                    asset_type = "MUTUAL_FUND"
                elif any(k in row_text for k in ["f&o", "future", "option", "derivatives"]):
                    asset_type = "FNO"
                elif any(k in row_text for k in ["property", "real estate", "land", "building"]):
                    asset_type = "PROPERTY"
                else:
                    asset_type = "LISTED_EQUITY"

                values = {
                    "asset_type": asset_type,
                    "description": symbol,
                    "acquisition_date": buy_date,
                    "transfer_date": sell_date,
                    "sale_consideration": str(sale) if sale is not None else None,
                    "cost_of_acquisition": str(cost) if cost is not None else None,
                    "transfer_expenses": str(expenses),
                    "broker": sheet_name,
                    "original_row": idx + 2,  # +2 for header and 1-based indexing
                }

                claim = ExtractedClaim(
                    field_code="CAPITAL_GAIN.TRANSACTION",
                    value_type="capital_gain",
                    value=values,
                    source=SourceLocation(original_text=symbol),
                    confidence=Decimal("0.92"),
                )
                claims.append(claim)

        wb.close()

        if not claims:
            warnings.append("No capital gains transactions could be extracted. Check format.")

        return AdapterResult(
            self.code, self.version, self.document_type,
            claims=claims, warnings=warnings,
            metadata={"transactions_found": len(claims)}
        )


class ExcelBankStatementAdapter(DocumentAdapter):
    """Adapter for bank statements in Excel format."""
    code = "EXCEL_BANK_STATEMENT"
    version = "1.0.0"
    document_type = "BANK_STATEMENT"

    def supports(self, filename: str, mime_type: str, content: bytes) -> Decimal:
        if not HAS_OPENPYXL:
            return Decimal("0")
        name = filename.lower()
        if not (name.endswith(".xlsx") or name.endswith(".xls")):
            return Decimal("0")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text = " ".join(str(cell.value or "") for row in sheet.iter_rows(max_row=20, max_col=5) for cell in row)
                if any(k in text.lower() for k in ["bank", "statement", "account", "transaction", "balance", "debit", "credit"]):
                    return Decimal("0.85")
            wb.close()
        except Exception:
            pass
        return Decimal("0.20")

    def extract(self, filename: str, mime_type: str, content: bytes) -> AdapterResult:
        if not HAS_OPENPYXL:
            return AdapterResult(
                self.code, self.version, self.document_type,
                warnings=["openpyxl not installed. Please install: pip install openpyxl"]
            )

        claims: list[ExtractedClaim] = []
        warnings: list[str] = []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

            total_credit = Decimal("0")
            total_debit = Decimal("0")
            transaction_count = 0

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = _read_excel_sheet(sheet)

                for row in rows:
                    # Try to find amount columns
                    for key in ["Amount", "Debit", "Credit", "Balance", "Value"]:
                        val = _parse_decimal_excel(row.get(key))
                        if val is not None:
                            if val > 0:
                                total_credit += val
                            else:
                                total_debit += abs(val)
                            transaction_count += 1
                            break

            wb.close()

            if transaction_count > 0:
                claims.append(ExtractedClaim(
                    field_code="OTHER_INCOME.BANK_INTEREST.TOTAL",
                    value_type="money",
                    value={"amount": str(total_credit), "transaction_count": transaction_count},
                    source=SourceLocation(original_text=f"Excel statement with {transaction_count} transactions"),
                    confidence=Decimal("0.75"),
                ))
            else:
                warnings.append("Could not extract amounts from Excel. Please use CSV format if possible.")

        except Exception as e:
            warnings.append(f"Error reading Excel file: {str(e)}")

        return AdapterResult(
            self.code, self.version, self.document_type,
            claims=claims, warnings=warnings
        )
